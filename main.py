import argparse
import csv
import datetime
import mechanicalsoup
from openpyxl import load_workbook
import os
#from prompt_toolkit.shortcuts import radiolist_dialog, input_dialog, yes_no_dialog
from streamlit.web import cli as stcli
import sys

from database import RefereeDbCockroach
from refWebSites import MySoccerLeague

db = RefereeDbCockroach()

def getRealTimeCurrentRefAssignments(br: mechanicalsoup.stateful_browser.StatefulBrowser) -> dict:
    """
    Log into the MySoccerLeague website and pull all assignments for the weekend"""
    site = MySoccerLeague(br)
    assignments = site.getAssignments()
    return assignments


def getPastAssignments(br: mechanicalsoup.stateful_browser.StatefulBrowser, d: datetime.date) -> dict:
    site = MySoccerLeague(br)
    site.setSpecificDate(d)
    return site.getAssignments()


def getAllRefereesFromSite(br: mechanicalsoup.stateful_browser.StatefulBrowser) -> list:
    site = MySoccerLeague(br)
    return site.getAllReferees()


# def getNewRefereesFromGoogleSheet(filename: str) -> dict:
#     results = {}

#     # Open the Workbook
#     workbook = load_workbook(filename = filename)
#     sheet = workbook.active

#     # Iterate the rows
#     for row in sheet.iter_rows(min_row = 2):

#         if row[0].value == None:
#             break

#         name = row[2].value.lower().strip()
#         lastName, firstName = name.split(',')

#         year = row[7].value.year

#         if lastName == '':
#             break

#         results[f'{firstName.strip()} {lastName.strip()}'] = year

#     return results


def getAllReferees() -> dict:
    """
    Organize data from Dianne's MSL dump.
    """
    results = {}

    ##  can we get rid of this step?  Get straight from web site
    # Open the Workbook
    workbook = load_workbook(filename = "newRefs/NewRefs10182022.xlsx")
    sheet = workbook.active

    # Iterate the rows
    index = 0
    for row in sheet.iter_rows(min_row = 2):

        lastName = row[7].value.lower().strip()
        firstName = row[6].value.lower().strip()
        id = row[2].value

        if index > 30 and lastName == '':
            break

        results[f'{firstName} {lastName}'] = id
        index += 1
    return results


def getRefsAlreadyMentored() -> dict:
    """
    Pull the names of all referees already mentored this season
    """
    retVal = []
    sessions = db.getMentoringSessions()

    for session in sessions:
        retVal.append(f'{session[0].strip().lower()} {session[1].strip().lower()}')

    return retVal


def getRefsAlreadyMentoredOld() -> list:
    """
    Pull the names of all referees already mentored this season
    """
    retVal = []
    for x in os.listdir('reports/fall2022'):
        # these filenames are firstname_lastname.txt
        if x.endswith(".txt"):
            filename = x.split('.')
            refName = filename[0].split('_')
            retVal.append(f'{refName[0].strip().lower()} {refName[1].strip().lower()}')
            # workbook = load_workbook(filename = f'reports/{x}')
            # sheet = workbook.active

            # # Iterate the rows
            # for i, row in enumerate(sheet.iter_rows(min_row = 2)):
            #     center = row[5].value

            #     if center == '' or center is None:
            #         break

            #     center = center.lower().strip()
            #     retVal.append(center)

            #     ar1 = row[6].value
            #     if ar1 is not None and ar1 != 'n/a':
            #         ar1 = ar1.lower().strip()
            #         retVal.append(ar1)
            #     ar2 = row[7].value
            #     if ar2 is not None and ar1 != 'n/a':
            #         ar2 = ar2.lower().strip()
            #         retVal.append(ar2)

    return retVal


def printout(currentu: list, newRefs: list, mentored: list, skip: bool, report: bool=False) -> None:
    current = {}
    for c in sorted(currentu):
        current[c] = currentu[c]

    if skip:
        print("Current Schedule of New Referees never mentored:")
    elif not report:
        print("Current Schedule of New Referees:")

    for field, details in current.items():
        fieldsOnce = False
        for game in details:

            center = details[game]['Center'].lower()
            ar1 = details[game]['AR1'].lower()
            ar2 = details[game]['AR2'].lower()

            if center not in newRefs and ar1 not in newRefs and ar2 not in newRefs:
                continue

            cmarker = '**' if center in mentored else ''
            a1marker = '**' if ar1 in mentored else ''
            a2marker = '**' if ar2 in mentored else ''

            # only print out refs that are new and not previously mentored
            if skip and (center not in newRefs or cmarker == '**') \
                and (ar1 not in newRefs or a1marker == '**') \
                     and (ar2 not in newRefs or a2marker == '**'):
                continue

            if not fieldsOnce:
                print("")
                print(f'Field: {field}')
                fieldsOnce = True

            date = details[game]['date']
            gameTime = details[game]['gameTime']
            age = details[game]['age']
            level = details[game]['level']

            print(f'\tID: {game}, Date: {date}, Time: {gameTime}, Age: {age}, Level: {level}')

            if not (skip and cmarker == '**'):
                if center in newRefs:
            #if center in newRefs and cmarker != '**':
                    print(f'\t\tNew Ref at Center: {center.title()}{cmarker}')
            if not (skip and a1marker == '**'):
                if ar1 in newRefs:
            #if ar1 in newRefs and a1marker != '**':
                    print(f'\t\tNew Ref at AR1: {ar1.title()}{a1marker}')
            if not (skip and a2marker == '**'):
                if ar2 in newRefs:
            #if ar2 in newRefs and a2marker != '**':
                    print(f'\t\tNew Ref at AR2: {ar2.title()}{a2marker}')
    if not skip and not report:
        print("** Referee has already had a mentor")
    print("")


def run(skip: bool) -> None:

    """
    Verify new referees have the same first and last name in MSL.
    """
    newRefs = db.getNewReferees(2023)
    br = mechanicalsoup.StatefulBrowser(soup_config={ 'features': 'lxml'})
    br.addheaders = [('User-agent', 'Chrome')]

    allRefs = getAllRefereesFromSite(br)

    for ref in newRefs:
        if ref not in allRefs:
            print (f'Referee: {ref} not in MSL, check name spelling')

    """
    Gather the new referee data from Dianne and correlate with current week's assignment.
    """

    # get this week's current assignments
    current = getRealTimeCurrentRefAssignments(br)

    # get list of already mentored referees
    mentored = getRefsAlreadyMentored()

    #printout(current, newRefs)
    printout(current, newRefs, mentored, skip)
    printout(current, newRefs, mentored, not skip)


def produceReport() -> None:

    newRefs = getNewRefereesFromGoogleSheet("newRefs/googlesheet.xlsx")

    br = mechanicalsoup.StatefulBrowser(soup_config={ 'features': 'lxml'})
    br.addheaders = [('User-agent', 'Chrome')]
    site = MySoccerLeague(br)
    dates = site.getAllDatesForSeason()
    today = datetime.date.today()
    todaysDate = datetime.datetime.combine(today, datetime.time(0, 0))
    assignments = []
    for date in dates:
        dt = datetime.datetime.strptime(date, "%A, %B %d, %Y")
        if dt < todaysDate:
            if dt.weekday() == 5: # Saturday
                site.setSpecificDate(dt)
                assignments.append(site.getAssignments())
    for a in assignments:
        printout(a, newRefs, [], False, True)


if __name__ == "__main__":

    #sys.argv = ['streamlit', 'run', '--server.port', '443', 'ui.py']
    #stcli.main()

    from uiData import getAllData
    x = getAllData()

    parser = argparse.ArgumentParser()
    parser.add_argument('-r', '--report', dest='report', action='store_true', help='run season report')

    args = parser.parse_args()

    if args.report:
        produceReport()
    else:
        run(skip = True)
